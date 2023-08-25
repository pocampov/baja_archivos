# Aplicación para la generación de tablas para el Informe de
# análisis de oferta de las RIPPS
#
import datetime

import tkinter as tk
from tkinter import PhotoImage
from tkinter import ttk
from tkinter import filedialog

from tqdm import tqdm

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from urllib.parse import urljoin
import os
import sys
import requests
from busca_version import busca_version

# variables globales
version = "1.3"
archivo_configuracion = "config_tablas.txt"
download_url = "https://misejecutables.000webhostapp.com" # Ubicación de nuevas versiones
# === Funciones para auto-actualizar el programa
def copia_nueva_version():
    nombre_programa = "Tablas_para_RIPSS.exe"
    # Si el archivo que se ejecuta se llama nueva_version, copia este archivo en baja_archivos
    if os.path.basename(sys.argv[0]) == "nueva_version.exe":
        try:
            os.replace("nueva_version.exe", nombre_programa)
        except Exception as e:
            print("Error al reemplazar el archivo:", e)
            sys.exit(1)
        print("¡Actualización completada! Reiniciando el programa...")
        # Reiniciar el programa con el nombre original
        os.system("start "+nombre_programa)
        sys.exit()


def download_latest_version(url, dir):
    return busca_version(url,dir)

def restart_program():
    python = sys.executable
    os.execl(python, python, *sys.argv)

def extrae_ultima_version(url, filename):
    new_filename = "nueva_version.exe"
    with open(new_filename, 'wb') as f:
        url = urljoin(url, filename)
        print(url)
        response = requests.get(url, stream=True)
        total_length = response.headers.get('content-length')
        print(total_length)
        if total_length is None:
            f.write(response.content)
        else:
            dl = 0
            total_length = int(total_length)
            print("Directorio de trabajo actual:", os.getcwd())
            progress_bar = tqdm(total=total_length, unit='B', unit_scale=True, ncols=80, miniters=1)
            for data in response.iter_content(chunk_size=4096):
                dl += len(data)
                f.write(data)
                progress_bar.update(len(data))
            progress_bar.close()
def boton_actualizar():
    posx = 4
    posy = 1
    boton_actualiza = tk.Button(root, text="Actualizar versiónn", command=actualiza)
    boton_actualiza.grid(row=posx, column=posy,columnspan=1, pady=1)
def update_program():
    global version, download_url
    current_version = version  # Obtener la versión actual de tu programa
    last_version = download_latest_version(download_url,"tpr")
    if not last_version:
        last_version = version
    print("Version actual: "+ current_version + " Ultima versión: " + str(last_version))
    if current_version < str(last_version):
        boton_actualizar()
        return
    else:
        print("No hay actualizaciones disponibles.")
def actualiza():
    new_executable = "tpr/Tablas_para_RIPSS.exe"
    # Descargar el nuevo ejecutable
    global download_url
        
    extrae_ultima_version(download_url, new_executable)

    # Cerrar el archivo en ejecución ("baja_archivos.exe")
    # Reiniciar el programa con el nombre original
    print("¡Archivo descargao! Reiniciando el programa...")
    os.system("start nueva_version.exe")
    sys.exit()

# === Fin funciones de auto-actualización


def carga_libros(sel):
    global sheet_ripss
    global sheet_prestadores
    global sheet_capacidad
    
    # Cargar los libros de trabajo
    if sel == 1:
        global archivo_RIPSS, workbook_ripss, archivo_destino
        workbook_ripss = load_workbook(filename=archivo_destino, read_only=True)  
        sheet_name_ripss = "BD_Depurada"
        sheet_ripss = workbook_ripss[sheet_name_ripss]
    if sel == 2: 
        global archivo_prestadores, archivo_origen
        workbook_prestadores = load_workbook(filename=archivo_origen, read_only=True)
        sheet_name_prestadores = "Prestadores (3)"    
        sheet_prestadores = workbook_prestadores[sheet_name_prestadores]
    if sel == 3: 
        global archivo_codigos_capacidad
        workbook_capacidad = load_workbook(filename=archivo_codigos_capacidad, read_only=True)
        sheet_name_capacidad = "Hoja1"    
        sheet_capacidad = workbook_capacidad[sheet_name_capacidad]

def selecciona_carpeta(sel):
    global archivo_prestadores, archivo_RIPSS
    global archivo_destino, archivo_origen, archivo_codigos_capacidad
    ruta_seleccionada = os.path.join(os.environ['USERPROFILE'], 'Documents')
    
    if ruta_seleccionada:
        if sel == 1:
            ruta_seleccionada = recupera_parametro("archivo_destino")
            directorio, nombre_archivo = os.path.split(ruta_seleccionada)
            ruta_seleccionada = filedialog.askopenfilename(title="Ubicación de la fuente de RIPSS", initialdir=directorio,filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")) )
            entry.delete(0, tk.END)  # Limpiar el contenido actual del Entry
            archivo_destino = ruta_seleccionada
            entry.insert(0, ruta_seleccionada)
            archivo_RIPSS = ruta_seleccionada
            actualizar_combobox()
        if sel == 2:
            ruta_seleccionada = recupera_parametro("archivo_origen")
            directorio, nombre_archivo = os.path.split(ruta_seleccionada)
            ruta_seleccionada = filedialog.askopenfilename(title="Ubicación de la fuente de Prestadores", initialdir=directorio,filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")) )
            entry2.delete(0, tk.END)  # Limpiar el contenido actual del Entry
            archivo_origen = ruta_seleccionada
            print(ruta_seleccionada)
            entry2.insert(0, ruta_seleccionada)
            archivo_prestadores = ruta_seleccionada
        if sel == 3:
            ruta_seleccionada = recupera_parametro("archivo_codigos_capacidad")
            directorio, nombre_archivo = os.path.split(ruta_seleccionada)
            ruta_seleccionada = filedialog.askopenfilename(title="Ubicación de la fuente de códigos capacidad instalada", initialdir=directorio,filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")) )
            entry3.delete(0, tk.END)  # Limpiar el contenido actual del Entry
            archivo_codigos_capacidad = ruta_seleccionada
            print(ruta_seleccionada)
            entry3.insert(0, ruta_seleccionada)
            archivo_codigos_capacidad = ruta_seleccionada
    
def actualiza_ripss():
    global sheet_ripss, sheet_prestadores, workbook_ripss
    global archivo_origen, archivo_destino
    asigna_parametro("archivo_origen")
    asigna_parametro("archivo_destino")
    carga_libros(1) # Excel de destino
    carga_libros(2) # Excel de origen de datos
    # Crea nueva hoja con permisos de escritura y copia sheet_ripss
    workbook_writable = Workbook()  # Aquí se crea el archivo 
    sheet_writable = workbook_writable.active
    # Copiar los datos de la hoja sheet_ripss de solo lectura a la hoja en el archivo editable
    for row in sheet_ripss.iter_rows(values_only=True):
        sheet_writable.append(row)
    
    # Obtener índices de columnas basados en encabezados
    col_index_ripss_codigo_prestador = 1
    col_index_prestadores_codigo_habilitacion = 2 
    
    # Crear un diccionario para almacenar los valores deseados
    valores_deseados = {}
    modificar = []
    # Iterar a través de las filas en archivo_prestadores y almacenar los valores en el diccionario
    for row_prestadores in sheet_prestadores.iter_rows(min_row=2, values_only=True):
        codigo_habilitacion_prestadores = row_prestadores[col_index_prestadores_codigo_habilitacion]
        valor_deseado = [row_prestadores[30],row_prestadores[31],row_prestadores[32],
                         row_prestadores[33],row_prestadores[34],row_prestadores[35],
                         row_prestadores[36],row_prestadores[37],row_prestadores[38]
                         ]
        valores_deseados[str(codigo_habilitacion_prestadores)] = valor_deseado
    
    # Iterar a través de las filas en archivo_RIPSS y actualizar las celdas
    for contador_fila_ripss, row_ripss in enumerate(sheet_writable.iter_rows(min_row=2, values_only=True), start=2):
        codigo_prestador_ripss = row_ripss[col_index_ripss_codigo_prestador]
        valor_deseado = valores_deseados.get(str(codigo_prestador_ripss))
        #print(contador_fila_ripss)
        if valor_deseado is not None:
            sheet_writable.cell(row=contador_fila_ripss, column=22).value = valor_deseado[0]
            sheet_writable.cell(row=contador_fila_ripss, column=23).value = valor_deseado[1]
            sheet_writable.cell(row=contador_fila_ripss, column=24).value = valor_deseado[2]
            sheet_writable.cell(row=contador_fila_ripss, column=25).value = valor_deseado[3]
            sheet_writable.cell(row=contador_fila_ripss, column=26).value = valor_deseado[4]
            sheet_writable.cell(row=contador_fila_ripss, column=27).value = valor_deseado[5]
            sheet_writable.cell(row=contador_fila_ripss, column=28).value = valor_deseado[6]
            sheet_writable.cell(row=contador_fila_ripss, column=29).value = valor_deseado[7]
            sheet_writable.cell(row=contador_fila_ripss, column=30).value = valor_deseado[8]

    # Guardar los cambios en archivo_RIPSS
    workbook_writable.save('archivo_RIPSS_actualizado.xlsx')

def crea_tabla_1():
    hint.config(text="Elaborando Tabla 1 ...")
    root.update()
    global sheet_ripss  #, archivo_RIPSS
    carga_libros(1)
    tipo_prestador = []
    prestadores  = []
    no_prestadores = []
    eps = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True):
        tipo_prestador = str(row[15])
        eps.append(row[0])
        if tipo_prestador == "Instituciones Prestadoras de Servicios de Salud - IPS":
            prestadores.append(str(row[1]))
        else:
            no_prestadores.append(str(row[1]))
    # Resultados
    total_prestadores = len(set(prestadores))
    total_otros = len(set(no_prestadores))
    servicios_no_prestadores = len(no_prestadores)
    servicios_prestadores = len(prestadores)
    # Dibuja la tabla
    data = [
    ["PROVEEDORES",	"CANTIDAD",	"SERVICIOS CONTRATADOS"],
    ["Prestadores de servicios de salud",total_prestadores,servicios_prestadores],
    ["Otros Proveedores",total_otros, servicios_no_prestadores],
    ["Total", total_prestadores + total_otros, servicios_prestadores + servicios_no_prestadores]
    ]
    return data

def crea_tabla_2():
    hint.config(text="Elaborando Tabla 2 ...")
    root.update()
    global sheet_ripss  #, archivo_RIPSS
    carga_libros(1)

    # Obtiene información
    prestadores = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True):
        tipo_prestador = str(row[15])
        codigo = str(row[1])
        prestadores.append((codigo, tipo_prestador))
    total_IPS = len(set((x,y) for x,y in prestadores if y==str("Instituciones Prestadoras de Servicios de Salud - IPS")))
    total_diferente = len(set((x,y) for x,y in prestadores if y==str("Objeto Social Diferente a la Prestación de Servicios de Salud")))
    total_otros = len(set((x,y) for x,y in prestadores if y==str("Otros prestadores de servicios")))
    total_profesional = len(set((x,y) for x,y in prestadores if y==str("Profesional Independiente")))
    total_transporte = len(set((x,y) for x,y in prestadores if y==str("Transporte Especial de Pacientes")))
    total = total_IPS + total_diferente + total_otros + total_profesional + total_transporte
    # Crea la matriz (tabla)
    titulos = ["TIPO PRESTADOR","PRESTADORES","%"]
    data = [
            titulos,
            ["IPS",total_IPS,round(total_IPS*100/total,1) ],
            ["OBJETO SOCIAL DIFERENTE",total_diferente,round(total_diferente*100/total,1)],
            ["OTROS PRESTADORES DE SERVICIO",total_otros,round(total_otros*100/total,1)],
            ["PROFESIONAL INDEPENDIENTE",total_profesional,round(total_profesional*100/total,1)],
            ["TRANSPORTE ESPECIAL DE PACIENTES",total_transporte,round(total_transporte*100/total,1)],
            ["TOTAL",total,100]
        ]
    return data

def crea_tabla_3():
    hint.config(text="Elaborando Tabla 3 ...")
    root.update()
    global sheet_ripss  #, archivo_RIPSS
    carga_libros(1)

    # Obtiene información
    prestadores = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True):
        if str(row[15]) == "Instituciones Prestadoras de Servicios de Salud - IPS":
            eps = str(row[0])
            codigo = str(row[1])
            prestadores.append((eps,codigo))
    Eps = set(x for x,y in prestadores)
    total_eps = len(Eps)
    
    # Crea la matriz (tabla)
    data = []
    titulos = ["EAPB",	"NÚMERO IPS","AFILIADOS EAPB","PROPORCIÓN por cada 10.000 afiliados"]
    data.append(titulos)
    for nombre_eps in sorted(list(Eps)):
        total_ips = len(set( y for x,y in prestadores if x == nombre_eps))
        data.append([nombre_eps,total_ips,"",""])
    return data

def crea_tabla_4():
    hint.config(text="Elaborando Tabla 4 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    global sheet_prestadores # archivo con prestadores
    carga_libros(1)
    carga_libros(2)
    # Crear un diccionario codigo_ips: array geolocalización
    diccionario_ips = {}
        # Iterar a través de las filas en archivo_prestadores y almacenar los valores en el diccionario
    for row_prestadores in sheet_prestadores.iter_rows(min_row=2, values_only=True):
        codigo_habilitacion_prestadores = str(row_prestadores[2])
        valor_deseado = row_prestadores[0:39]
        diccionario_ips[str(codigo_habilitacion_prestadores)] = valor_deseado
    # Obtiene información
    prestadores = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True):
            codigo = str(row[1])
            nom_loc = diccionario_ips.get(codigo, False)
            if nom_loc:
                nombre_localidad = nom_loc[32]
            else:
                nombre_localidad = "Sin localizar"
            prestadores.append((codigo,nombre_localidad))
    localidades = sorted(set(y for x,y in prestadores))

    ips_loc = set(prestadores)
    # Crea la matriz (tabla)
    data = []
    titulos = ["LOCALIDAD", "PRESTADORES",	"%"]
    data.append(titulos)
    tot = len(ips_loc)
    for localidad in localidades:
        total_ips = len(set( x for x,y in ips_loc if y == localidad))
        data.append([localidad,total_ips,round((total_ips/tot)*100,2)])
    data.append(["TOTAL",tot,100])
    return data

def crea_tabla_5():
    hint.config(text="Elaborando Tabla 5 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    carga_libros(1)
    #carga_libros(2)

    # Obtiene información
    grupo_servicio = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        grupo_servicio.append(str(row[16]))
    servicios = set(grupo_servicio)
    
    # Crea la matriz (tabla)
    data = []
    titulos = ["GRUPO SERVICIOS","TOTAL","%"]
    data.append(titulos)
    tot = len(grupo_servicio)
    for grupo in servicios:
        total_grupo = grupo_servicio.count(grupo)
        data.append([grupo,total_grupo,round((total_grupo/tot)*100,2)])
    data.append(["TOTAL",tot,100])
    return data

def crea_tabla_6():
    hint.config(text="Elaborando Tabla 6 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    carga_libros(1)
    #carga_libros(2)

    # Obtiene información
    red_general = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        eps = str(row[0])
        if row[7] == None:
            red_general.append( (eps, "NONE") )
        else:
            red_general.append( (eps, str(row[7])) )
    nombres_eps = set(x.strip().upper() for x,y in red_general)
    tipos = list(set(y.strip().upper() for x,y in red_general))
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["EAPB"] + [x for x in tipos] + ["TOTAL"]
    data.append(titulo1)
    total = {}

    for nombre in sorted(list(nombres_eps)):
        for tipo in tipos:
            total[tipo] = red_general.count((nombre,tipo))
        tot = sum(total.values())
        data.append([nombre] +[total[x] for x in tipos] + [tot])
    
    suma_columnas = [0] * len(data[0])
    for fila in data[1:]:
        for i in range(1,len(fila)):
            suma_columnas[i] += fila[i]
    
    data.append(["TOTAL"] + suma_columnas[1:])
    return data

def crea_tabla_7():
    hint.config(text="Elaborando Tabla 7 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    carga_libros(1)
    #carga_libros(2)

    # Obtiene información
    red_general = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        eps = str(row[0])
        if row[8] == None:
            red_general.append( (eps, "NONE") )
        else:
            red_general.append( (eps, str(row[8]) ))
    nombres_eps = set(x.strip().upper() for x,y in red_general)
    tipos = list(set(y.strip().upper() for x,y in red_general))
    print(tipos)
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["EAPB"] + [x.upper() for x in tipos] + ["TOTAL"]
    data.append(titulo1)
    total = {}
    for nombre in sorted(list(nombres_eps)):
        for tipo in tipos:
            total[tipo] = red_general.count((nombre,tipo))
        tot = sum(total.values())
        data.append([nombre] +[total[x] for x in tipos] + [tot])
    
    suma_columnas = [0] * len(data[0])
    for fila in data[1:]:
        for i in range(1,len(fila)):
            suma_columnas[i] += fila[i]
    
    data.append(["TOTAL"] + suma_columnas[1:])
    return data

def crea_tabla_8():
    hint.config(text="Elaborando Tabla 8 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    carga_libros(1)
    #carga_libros(2)

    # Obtiene información
    red_general = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        eps = str(row[0])
        if row[9] == None:
            red_general.append( (eps, "NONE") )
        else:
            red_general.append( (eps, str(row[9])) )
    nombres_eps = set(x.strip().upper() for x,y in red_general)
    tipos = list(set(y.strip().upper() for x,y in red_general))
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["EAPB"] + [x for x in tipos] + ["TOTAL"]
    data.append(titulo1)
    total = {}
    for nombre in sorted(list(nombres_eps)):
        for tipo in tipos:
            total[tipo] = red_general.count((nombre,tipo))
        tot = sum(total.values())
        data.append([nombre] +[total[x] for x in tipos] + [tot])
    
    suma_columnas = [0] * len(data[0])
    for fila in data[1:]:
        for i in range(1,len(fila)):
            suma_columnas[i] += fila[i]
    
    data.append(["TOTAL"] + suma_columnas[1:])
    return data

def crea_tabla_9():
    hint.config(text="Elaborando Tabla 9 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    carga_libros(1)
    #carga_libros(2)

    # Obtiene información
    red_general = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        eps = str(row[0])
        if row[10] == None:
            red_general.append( (eps, "NONE") )
        else:
            red_general.append( (eps, str(row[10])) )
    nombres_eps = set(x.strip() for x,y in red_general)
    tipos = list(set(y.strip().upper() for x,y in red_general))
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["EAPB"] + [x for x in tipos] + ["TOTAL"]
    data.append(titulo1)
    total = {}
    for nombre in sorted(list(nombres_eps)):
        for tipo in tipos:
            total[tipo] = red_general.count((nombre,tipo))
        tot = sum(total.values())
        data.append([nombre] +[total[x] for x in tipos] + [tot])
    
    suma_columnas = [0] * len(data[0])
    for fila in data[1:]:
        for i in range(1,len(fila)):
            suma_columnas[i] += fila[i]
    
    data.append(["TOTAL"] + suma_columnas[1:])
    return data

def crea_tabla_10(nombre_eps):
    hint.config(text="Elaborando Tabla 10 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    carga_libros(1)
    #carga_libros(2)

    # Obtiene información
    if len(combobox_eps.cget("values")) <= 1:
           print("Debe seleccionar una fuente de datos de RIIPS")
           hint.config(text="Debe seleccionar una fuente de datos de RIIPS")
    red_general = []
    primario = []
    complementario = []
    mixto = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        if row[0] == nombre_eps:
            cod_prestador = str(row[1]).strip()
            cod_sede = str(row[3]).strip()
            servicio = str(row[5]).strip()
            red_general = str(row[7]).strip()
            red_oncologica = str(row[8]).strip()
            red_urgencias = str(row[9]).strip()
            red_alto_no_oncologica = str(row[10]).strip()

            if red_general == "PRIMARIO":
                primario.append(("Red General",cod_prestador,cod_sede,servicio))
            if red_general == "COMPLEMENTARIO":
                complementario.append(("Red General",cod_prestador,cod_sede,servicio))
            if red_general == "MIXTO":
                mixto.append(("Red General",cod_prestador,cod_sede,servicio))
            if red_oncologica == "PRIMARIO":
                primario.append(("Red Oncologica",cod_prestador,cod_sede,servicio))
            if red_oncologica == "COMPLEMENTARIO":
                complementario.append(("Red Oncologica",cod_prestador,cod_sede,servicio))
            if red_oncologica == "MIXTO":
                mixto.append(("Red Oncologica",cod_prestador,cod_sede,servicio))
            if red_urgencias == "PRIMARIO":
                primario.append(("Red Urgencias",cod_prestador,cod_sede,servicio))
            if red_urgencias == "COMPLEMENTARIO":
                complementario.append(("Red Urgencias",cod_prestador,cod_sede,servicio))
            if red_urgencias == "MIXTO":
                mixto.append(("Red Urgencias",cod_prestador,cod_sede,servicio))
            if red_alto_no_oncologica == "PRIMARIO":
                primario.append(("Red Alto Costo no Oncologica",cod_prestador,cod_sede,servicio))
            if red_alto_no_oncologica == "COMPLEMENTARIO":
                complementario.append(("Red Alto Costo no Oncologica",cod_prestador,cod_sede,servicio))
            if red_alto_no_oncologica == "MIXTO":
                mixto.append(("Red Alto Costo no Oncologica",cod_prestador,cod_sede,servicio))
    redes = list(set([x for (x,y,z,w) in complementario]+[x for (x,y,z,w) in mixto]))
    print(redes)
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["Tipo Red","Prestadores","Sedes","Servicios",
               "Prestadores ","Sedes ","Servicios ",
               "Prestadores  ","Sedes  ","Servicios  "
               ]
    data.append(titulo1)
    for nombre in redes:
        tot_prestadores = len(list(set([(x,y) for (x,y,z,w) in primario if x == nombre])))
        tot_sede = len(set([(x,y,z) for x,y,z,w in primario if x == nombre]))
        tot_servicios = len(set([(x,w) for x,y,z,w in primario if x == nombre]))
        
        tot_prestadores1 = len(list(set([(x,y) for (x,y,z,w) in complementario if x == nombre])))
        tot_sede1 = len(set([(x,y,z) for x,y,z,w in complementario if x == nombre]))
        tot_servicios1 = len(set([(x,w) for x,y,z,w in complementario if x == nombre]))
        
        tot_prestadores2 = len(list(set([(x,y) for (x,y,z,w) in mixto if x == nombre])))
        tot_sede2 = len(set([(x,y,z) for x,y,z,w in mixto if x == nombre]))
        tot_servicios2 = len(set([(x,w) for x,y,z,w in mixto if x == nombre]))

        data.append([nombre, tot_prestadores, tot_sede, tot_servicios, 
                     tot_prestadores1, tot_sede1, tot_servicios1,
                     tot_prestadores2, tot_sede2, tot_servicios2
                     ] )

    return data

def crea_tabla_11(nombre_eps):
    hint.config(text="Elaborando Tabla 11 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    global sheet_prestadores # archivo con prestadores
    carga_libros(1)
    carga_libros(2)
    # Crear un diccionario codigo_ips: array geolocalización
    diccionario_ips = {}
        # Iterar a través de las filas en archivo_prestadores y almacenar los valores en el diccionario
    for row_prestadores in sheet_prestadores.iter_rows(min_row=2, values_only=True):
        codigo_habilitacion_prestadores = str(row_prestadores[2])
        valor_deseado = row_prestadores[0:39]
        diccionario_ips[str(codigo_habilitacion_prestadores)] = valor_deseado
    # Obtiene información
    prestadores = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        if str(row[0]).strip() == nombre_eps.strip():
            codigo_ips = str(row[1]).strip()
            registro_localidad = diccionario_ips.get(codigo_ips,False)
            localidad = "Sin Localizar"
            if registro_localidad:
                localidad = registro_localidad[32]
            prestadores.append((codigo_ips,localidad))
    nombres_localidad = sorted(list(set([y for (x,y) in prestadores ])))
    prestadores = list(set(prestadores)) # Elimina repetidos
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["LOCALIDAD",
            "PRESTADORES",
            "%"
            ]
    data.append(titulo1)
    totalg = len(prestadores)
    suma_porc = 0
    for nombre in nombres_localidad:
        num_prestadores = len([x for (x,y) in prestadores if y == nombre])
        data.append([nombre, num_prestadores, round((num_prestadores*100)/totalg,2)])
        suma_porc += (num_prestadores*100)/totalg
    data.append(["TOTAL", totalg, round(suma_porc,2)])
    return data


def crea_tabla_12(nombre_eps):
    hint.config(text="Elaborando Tabla 12 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    global sheet_capacidad # Códigos de capacidad instalada
    carga_libros(1)
    """
    carga_libros(3)
    # Crear un diccionario codigo_ips: array geolocalización
    diccionario_grupo_servicios = {}
        # Iterar a través de las filas en archivo_prestadores y almacenar los valores en el diccionario
    for row_grupo in sheet_capacidad.iter_rows(min_row=2, values_only=True):
        nombre_grupo = str(row_grupo[0])
        servicio = str(row_grupo[1])
        diccionario_grupo_servicios[servicio.strip()] = nombre_grupo.upper()
    """
    # Obtiene información
    registros = []
    sin_agrupar = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        if str(row[0]).strip() == nombre_eps.strip():
            servicio = str(row[5]).strip()
            red_general = str(row[7]).strip()
            grupo = str(row[16]).strip()
            #grupo = diccionario_grupo_servicios.get(servicio,False)
            if not grupo:
                grupo = "Servicio sin agrupar"
                sin_agrupar.append(servicio)
            if red_general in ["PRIMARIO","COMPLEMENTARIO","MIXTO"]:
                registros.append((grupo, red_general))
    print("Servicios sin agrupar: ")
    for item in set(sin_agrupar):
        print(item)
    # Redefinición nombres_grupo
    # nombres_grupo = sorted(set(diccionario_grupo_servicios.values()))
    nombres_grupo = list(set([x for (x,y) in registros]))
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["GRUPO DE SERVICIOS",
            "PRIMARIO",
            "COMPLEMENTARIO",
            "MIXTO",
            "TOTAL GENERAL",
            "%"
            ]
    data.append(titulo1)
    totalg = len(registros)
    suma_porc = 0
    for nombre in nombres_grupo:
        tot_primario = len([(x,y) for (x,y) in registros if x == nombre and y == "PRIMARIO"])
        tot_complementario = len([(x,y) for (x,y) in registros if x == nombre and y == "COMPLEMENTARIO"])
        tot_mixto = len([(x,y) for (x,y) in registros if x == nombre and y == "MIXTO"])
        total = tot_primario + tot_complementario + tot_mixto
        porcent = round((total * 100)/ totalg, 2)
        suma_porc += (total * 100)/ totalg
        data.append([nombre, tot_primario, tot_complementario, tot_mixto, total, porcent])
    t_primario = len([(x,y) for (x,y) in registros if y == "PRIMARIO"]) 
    t_complementario = len([(x,y) for (x,y) in registros if y == "COMPLEMENTARIO"])
    t_mixto = len([(x,y) for (x,y) in registros if y == "MIXTO"])
    
    data.append(["TOTAL", t_primario, t_complementario, t_mixto,totalg,round(suma_porc,2) ] )

    return data

def crea_tabla_13(nombre_eps):
    hint.config(text="Elaborando Tabla 13 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    global sheet_capacidad # Códigos de capacidad instalada
    carga_libros(1)
    carga_libros(3)
    # Crear un diccionario codigo_ips: array geolocalización
    diccionario_grupo_servicios = {}
        # Iterar a través de las filas en archivo_prestadores y almacenar los valores en el diccionario
    for row_grupo in sheet_capacidad.iter_rows(min_row=2, values_only=True):
        nombre_grupo = str(row_grupo[0])
        servicio = str(row_grupo[1])
        diccionario_grupo_servicios[servicio.strip()] = nombre_grupo.upper()
    # Obtiene información
    registros = []
    sin_agrupar = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        if str(row[0]).strip() == nombre_eps.strip():
            servicio = str(row[5]).strip()
            red_urgencias = str(row[9]).strip()
            grupo = str(row[16]).strip()
            #grupo = diccionario_grupo_servicios.get(servicio,False)
            if not grupo:
                grupo = "Servicio sin agrupar"
                sin_agrupar.append(servicio)
            if red_urgencias in ["PRIMARIO","COMPLEMENTARIO","MIXTO"]:
                registros.append((grupo, red_urgencias))
    print("Servicios sin agrupar: ")
    for item in set(sin_agrupar):
        print(item)
    # Redefinición nombres_grupo
    # nombres_grupo = sorted(set(diccionario_grupo_servicios.values()))
    nombres_grupo = list(set([x for (x,y) in registros]))
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["GRUPO DE SERVICIOS",
            "PRIMARIO",
            "COMPLEMENTARIO",
            "MIXTO",
            "TOTAL GENERAL",
            "%"
            ]
    data.append(titulo1)
    totalg = len(registros)
    suma_porc = 0
    for nombre in nombres_grupo:
        tot_primario = len([(x,y) for (x,y) in registros if x == nombre and y == "PRIMARIO"])
        tot_complementario = len([(x,y) for (x,y) in registros if x == nombre and y == "COMPLEMENTARIO"])
        tot_mixto = len([(x,y) for (x,y) in registros if x == nombre and y == "MIXTO"])
        total = tot_primario + tot_complementario + tot_mixto
        porcent = round((total * 100)/ totalg, 2)
        suma_porc += (total * 100)/ totalg
        data.append([nombre, tot_primario, tot_complementario, tot_mixto, total, porcent])
    t_primario = len([(x,y) for (x,y) in registros if y == "PRIMARIO"]) 
    t_complementario = len([(x,y) for (x,y) in registros if y == "COMPLEMENTARIO"])
    t_mixto = len([(x,y) for (x,y) in registros if y == "MIXTO"])
    
    data.append(["TOTAL", t_primario, t_complementario, t_mixto,totalg,round(suma_porc,2) ] )

    return data

def crea_tabla_14(nombre_eps):
    hint.config(text="Elaborando Tabla 14 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    global sheet_capacidad # Códigos de capacidad instalada
    carga_libros(1)
    carga_libros(3)
    # Crear un diccionario codigo_ips: array geolocalización
    diccionario_grupo_servicios = {}
        # Iterar a través de las filas en archivo_prestadores y almacenar los valores en el diccionario
    for row_grupo in sheet_capacidad.iter_rows(min_row=2, values_only=True):
        nombre_grupo = str(row_grupo[0])
        servicio = str(row_grupo[1])
        diccionario_grupo_servicios[servicio.strip()] = nombre_grupo.upper()
    # Obtiene información
    registros = []
    no_esta = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        if str(row[0]).strip() == nombre_eps.strip():
            servicio = str(row[5]).strip()
            red_oncologica = str(row[8]).strip()
            #grupo = diccionario_grupo_servicios.get(servicio,False)
            grupo = str(row[16]).strip()
            if not grupo:
                grupo = "Servicio sin agrupar"
            if red_oncologica in ["PRIMARIO","COMPLEMENTARIO","MIXTO"]:
                registros.append((grupo, red_oncologica))
    # Redefinición nombres_grupo
    # nombres_grupo = sorted(set(diccionario_grupo_servicios.values()))
    nombres_grupo = list(set([x for (x,y) in registros]))
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["GRUPO DE SERVICIOS",
            "PRIMARIO",
            "COMPLEMENTARIO",
            "MIXTO",
            "TOTAL GENERAL",
            "%"
            ]
    data.append(titulo1)
    totalg = len(registros)
    suma_porc = 0
    for nombre in nombres_grupo:
        tot_primario = len([(x,y) for (x,y) in registros if x == nombre and y == "PRIMARIO"])
        tot_complementario = len([(x,y) for (x,y) in registros if x == nombre and y == "COMPLEMENTARIO"])
        tot_mixto = len([(x,y) for (x,y) in registros if x == nombre and y == "MIXTO"])
        total = tot_primario + tot_complementario + tot_mixto
        porcent = round((total * 100)/ totalg, 2)
        suma_porc += (total * 100)/ totalg
        data.append([nombre, tot_primario, tot_complementario, tot_mixto, total, porcent])
    t_primario = len([(x,y) for (x,y) in registros if y == "PRIMARIO"]) 
    t_complementario = len([(x,y) for (x,y) in registros if y == "COMPLEMENTARIO"])
    t_mixto = len([(x,y) for (x,y) in registros if y == "MIXTO"])
    
    data.append(["TOTAL", t_primario, t_complementario, t_mixto,totalg,round(suma_porc,2) ] )

    return data

def crea_tabla_15(nombre_eps):
    hint.config(text="Elaborando Tabla 15 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    global sheet_capacidad # Códigos de capacidad instalada
    carga_libros(1)
    carga_libros(3)
    # Crear un diccionario codigo_ips: array geolocalización
    diccionario_grupo_servicios = {}
        # Iterar a través de las filas en archivo_prestadores y almacenar los valores en el diccionario
    for row_grupo in sheet_capacidad.iter_rows(min_row=2, values_only=True):
        nombre_grupo = str(row_grupo[0])
        servicio = str(row_grupo[1])
        diccionario_grupo_servicios[servicio.strip()] = nombre_grupo.upper()
    # Obtiene información
    registros = []
    no_esta = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        if str(row[0]).strip() == nombre_eps.strip():
            servicio = str(row[5]).strip()
            red_alto = str(row[10]).strip()
            #grupo = diccionario_grupo_servicios.get(servicio,False)
            grupo = str(row[16]).strip()
            if not grupo:
                grupo = "Servicio sin agrupar"
            if red_alto in ["PRIMARIO","COMPLEMENTARIO","MIXTO"]:
                registros.append((grupo, red_alto))
    # Redefinición nombres_grupo
    # nombres_grupo = sorted(set(diccionario_grupo_servicios.values()))
    nombres_grupo = list(set([x for (x,y) in registros]))
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["GRUPO DE SERVICIOS",
            "PRIMARIO",
            "COMPLEMENTARIO",
            "MIXTO",
            "TOTAL GENERAL",
            "%"
            ]
    data.append(titulo1)
    totalg = len(registros)
    suma_porc = 0
    for nombre in nombres_grupo:
        tot_primario = len([(x,y) for (x,y) in registros if x == nombre and y == "PRIMARIO"])
        tot_complementario = len([(x,y) for (x,y) in registros if x == nombre and y == "COMPLEMENTARIO"])
        tot_mixto = len([(x,y) for (x,y) in registros if x == nombre and y == "MIXTO"])
        total = tot_primario + tot_complementario + tot_mixto
        porcent = round((total * 100)/ totalg, 2)
        suma_porc += (total * 100)/ totalg
        data.append([nombre, tot_primario, tot_complementario, tot_mixto, total, porcent])
    t_primario = len([(x,y) for (x,y) in registros if y == "PRIMARIO"]) 
    t_complementario = len([(x,y) for (x,y) in registros if y == "COMPLEMENTARIO"])
    t_mixto = len([(x,y) for (x,y) in registros if y == "MIXTO"])
    
    data.append(["TOTAL", t_primario, t_complementario, t_mixto,totalg,round(suma_porc,2) ] )

    return data

def crea_tabla_16(nombre_eps):
    hint.config(text="Elaborando Tabla 16 ...")
    root.update()
    global sheet_ripss  # archivo_RIPSS
    # global sheet_prestadores # archivo con prestadores
    global sheet_capacidad # Códigos de capacidad instalada
    carga_libros(1)
    carga_libros(3)
    # Crear un diccionario codigo_ips: array geolocalización
    diccionario_grupo_servicios = {}
        # Iterar a través de las filas en archivo_prestadores y almacenar los valores en el diccionario
    for row_grupo in sheet_capacidad.iter_rows(min_row=2, values_only=True):
        nombre_grupo = str(row_grupo[0])
        servicio = str(row_grupo[1])
        diccionario_grupo_servicios[servicio.strip()] = nombre_grupo.upper()
    # Obtiene información
    registros = []
    no_esta = []
    for row in sheet_ripss.iter_rows(min_row=2, values_only=True): 
        if str(row[0]).strip() == nombre_eps.strip():
            servicio = str(row[5]).strip()
            # grupo = diccionario_grupo_servicios.get(servicio,False)
            grupo = str(row[16]).strip()
            if not grupo:
                grupo = "Servicio sin agrupar"
            registros.append((grupo, servicio))
    # Redefinición nombres_grupo
    # nombres_grupo = sorted(set(diccionario_grupo_servicios.values()))
    nombres_grupo = list(set([x for (x,y) in registros]))
    # Crea la matriz (tabla)
    data = []
    titulo1 = ["GRUPO DE SERVICIOS",
            "CANTIDAD DE SERVICIOS",
            "OFERTA TEÓRICA",
            "TOTAL ATENCIONES",	
            "SUFICIENCIA/DÉFICIT"
            ]
    data.append(titulo1)
    totalg = 0
    suma_porc = 0
    for nombre in nombres_grupo:
        tot_servicio = len([(x,y) for (x,y) in registros if x == nombre])
        data.append([nombre, tot_servicio, "", "", ""])
        totalg += tot_servicio
    
    data.append(["TOTAL", totalg, "", "", "" ] )

    return data


def dibuja_tabla(tabla, nombre, wb, sheet, titulo2=[]):
    hint.config(text="Dibujando "+nombre)
    root.update()
    if titulo2 != []:
        sheet.append(titulo2)
    for row in tabla:
        sheet.append(row)
    # Crear una tabla a partir de los datos
    # Calcular las dimensiones adecuadas para la nueva tabla
    num_rows = len(tabla)
    num_columns = len(tabla[0])
    start_cell = sheet.cell(row=sheet.max_row - num_rows + 1, column=1)
    end_cell = sheet.cell(row=sheet.max_row, column=num_columns)
    table_ref = f"{start_cell.coordinate}:{end_cell.coordinate}"

    # Crear una tabla a partir de los datos con la referencia calculada
    tab = Table(displayName=nombre, ref=table_ref)

    # Agregar un estilo a la tabla
    style = TableStyleInfo(
        name="TableStyleMedium1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    tab.tableStyleInfo = style

    # Ajustar el ancho de las columnas al tamaño de los títulos
    for column in sheet.iter_cols(min_col=start_cell.column, max_col=end_cell.column):
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2) * 1.2  # Ajuste del ancho, 1.2 es un factor arbitrario
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Agregar la tabla a la hoja de cálculo
    sheet.add_table(tab)
    nombre1 = nombre.replace("_", " ")
    sheet.append([nombre1])
    sheet.append([])
def crea_libro():
    global wb
    wb = Workbook()
    nombre_libro = "Tablas para RIPSS"
    asigna_parametro("archivo_origen")
    asigna_parametro("archivo_destino")
    asigna_parametro("archivo_codigos_capacidad")
    crea_hoja_Distrital()
    crea_hoja_Analisis_EAPB()
    # Guardar el libro de Excel
    i = 0
    nombre_guardar = nombre_libro+".xlsx"
    while os.path.exists(nombre_guardar):
        i += 1
        nombre_guardar = nombre_libro+" ("+str(i)+")"+".xlsx"
    wb.save(nombre_guardar)
    hint.config(text="Se ha creado el archivo "+nombre_guardar)
    root.update()

def crea_hoja_Distrital():
    nombre_hoja = "1.Distritales"
    hint.config(text="En ejecución ...")
    root.update()
    # Crear una nueva hoja activa
    global wb
    sheet = wb.active
    sheet.title = nombre_hoja

    # Tabla 1
    if checkbox_tabla1.get():
        tabla = crea_tabla_1()
        dibuja_tabla(tabla,"Tabla_1", wb, sheet)
    
    #Tabla 2
    if checkbox_tabla2.get():
        tabla = crea_tabla_2()
        dibuja_tabla(tabla,"Tabla_2", wb, sheet)

    #Tabla 3
    if checkbox_tabla3.get():
        tabla = crea_tabla_3()
        dibuja_tabla(tabla,"Tabla_3", wb, sheet)

    #Tabla 4
    if checkbox_tabla4.get():
        tabla = crea_tabla_4()
        dibuja_tabla(tabla,"Tabla_4", wb, sheet)
    
    #Tabla 5
    if checkbox_tabla5.get():
        tabla = crea_tabla_5()
        dibuja_tabla(tabla,"Tabla_5", wb, sheet)

    #Tabla 6
    if checkbox_tabla6.get():
        tabla = crea_tabla_6()
        dibuja_tabla(tabla,"Tabla_6._Red_General", wb, sheet)
    
    #Tabla 7
    if checkbox_tabla7.get():
        tabla = crea_tabla_7()
        dibuja_tabla(tabla,"Tabla_7._Red_Oncológica", wb, sheet) 

    #Tabla 8
    if checkbox_tabla8.get():
        tabla = crea_tabla_8()
        dibuja_tabla(tabla,"Tabla_8._Red_Urgencias", wb, sheet)

    #Tabla 9
    if checkbox_tabla9.get():
        tabla = crea_tabla_9()
        dibuja_tabla(tabla,"Tabla_9._RED_ALTO_COSTO_NO_ONCOLÓGICA", wb, sheet)

    # Guardar la hoja de Excel
    hint.config(text="Se ha creado la hoja: "+nombre_hoja)
    root.update()

def crea_hoja_Analisis_EAPB():
    nombre_hoja = "1.Análisis_EAPB"
    hint.config(text="En ejecución ...")
    root.update()
    # Crear una nueva hoja activa
    global wb
    sheet1 = wb.create_sheet(title=nombre_hoja)

    # Tabla 10
    if checkbox_tabla10.get():
        nombre_eps = combobox_eps.get()
        tabla = crea_tabla_10(nombre_eps)
        dibuja_tabla(tabla,"Tabla_10", wb, sheet1,[nombre_eps,"PRIMARIO", "","","COMPLEMENTARIO","","","MIXTO"])
    
    # Tabla 11
    if checkbox_tabla11.get():
        nombre_eps = combobox_eps.get()
        tabla = crea_tabla_11(nombre_eps)
        dibuja_tabla(tabla,"Tabla_11", wb, sheet1,["",nombre_eps,""])
    
    # Tabla 12
    if checkbox_tabla12.get():
        nombre_eps = combobox_eps.get()
        tabla = crea_tabla_12(nombre_eps)
        dibuja_tabla(tabla,"Tabla_12", wb, sheet1,["Red General. "+nombre_eps,"","","","",""])
    
    # Tabla 13
    if checkbox_tabla13.get():
        nombre_eps = combobox_eps.get()
        tabla = crea_tabla_13(nombre_eps)
        dibuja_tabla(tabla,"Tabla_13", wb, sheet1,["Red de Urgencias. "+nombre_eps,"","","","",""])
    
    # Tabla 14
    if checkbox_tabla14.get():
        nombre_eps = combobox_eps.get()
        tabla = crea_tabla_14(nombre_eps)
        dibuja_tabla(tabla,"Tabla_14", wb, sheet1,["Red de Oncológica. "+nombre_eps,"","","","",""])
    
    # Tabla 15
    if checkbox_tabla15.get():
        nombre_eps = combobox_eps.get()
        tabla = crea_tabla_15(nombre_eps)
        dibuja_tabla(tabla,"Tabla_15", wb, sheet1,["Red de Alto Costo no Oncológica. "+nombre_eps,"","","","",""])
       
    # Tabla 16
    if checkbox_tabla16.get():
        nombre_eps = combobox_eps.get()
        tabla = crea_tabla_16(nombre_eps)
        dibuja_tabla(tabla,"Tabla_16", wb, sheet1,["Suficiencia Servicios. "+nombre_eps,"","","","",""])
       

    # Guardar la hoja de Excel
    hint.config(text="Se ha creado la hoja: "+nombre_hoja)
    root.update()
# Funciones para consultar y almacenar parámetros
def asigna_parametro(variable):
    global archivo_configuracion
    print("El archivo es: "+archivo_configuracion)
    try:
        # Intentamos abrir el archivo en modo lectura
        with open(archivo_configuracion, "r") as file:
            lines = file.readlines()
    except FileNotFoundError:
        # Si el archivo no existe, lo creamos y escribimos el parámetro con su valor
        with open(archivo_configuracion, "w") as file:
            file.write(f"{variable}={globals()[variable]}\n")
    else:
        # Buscamos la variable en las líneas del archivo
        for i, line in enumerate(lines):
            key, value = line.strip().split("=")
            if key == variable:
                # Cambiamos el valor de la variable en el archivo
                lines[i] = f"{variable}={globals()[variable]}\n"
                with open(archivo_configuracion, "w") as file:
                    file.writelines(lines)
                break
        else:
            # Si la variable no está en el archivo, la agregamos al final
            with open(archivo_configuracion, "a") as file:
                file.write(f"{variable}={globals()[variable]}\n")

def recupera_parametro(variable):
    global archivo_configuracion
    file_name = archivo_configuracion
    
    try:
        # Intentamos abrir el archivo en modo lectura
        with open(archivo_configuracion, "r") as file:
            lines = file.readlines()
    except FileNotFoundError:
        # Si el archivo no existe, devolvemos el valor original de la variable
        return globals()[variable]
    else:
        # Buscamos la variable en las líneas del archivo
        for line in lines:
            key, value = line.strip().split("=")
            if key == variable:
                return value
        # Si la variable no está en el archivo, devolvemos el valor original de la variable
        return globals()[variable]
def filtra(cadena):
    return ''.join(caracter for caracter in cadena if caracter.isalnum() or caracter.isspace() or caracter == "-" or caracter == "_")
def listar_eps():
    global sheet_ripss
    eps = []
    eps.append(" Seleccione una EPS ")
    try:
        carga_libros(1)
        for row in sheet_ripss.iter_rows(min_row=2, values_only=True):
            nombre_eps = str(row[0])
            eps.append(nombre_eps.strip().upper())
        eps = list(set(eps))
    except Exception as e:
        print(e)
    return sorted(eps)
def actualizar_combobox():
    eps = listar_eps()
    combobox_eps.config(values=eps)
    print("Actualizar combobox")
def info_log():
    username = os.getlogin()
    fecha = datetime.date.today()
    hora = datetime.datetime.now().time()


# Programa principal
# Define directorios
script_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
images_dir = os.path.join(script_dir, 'images')
# Configuración de la ventana principal
root = tk.Tk()
root.title("Tablas para informe RIPSS  Ver.("+version+")")
root.geometry("535x510")
root.configure(bg="white")
root.resizable(False, False)
logo_path = os.path.join(images_dir, 'Logo.ico')
root.iconbitmap(logo_path)
icon_path = os.path.join(images_dir, 'folder.png')
icon_image_folder = PhotoImage(file=icon_path)
# Paso 1
title_frame_archivos = tk.Label(root, text="Paso 1: Seleccione los archivos fuente",font=("Calibri", 12),bg="white", fg="#29669b")
title_frame_archivos.grid(row=0, column=0, columnspan=5,padx=5,sticky="w")
# Frame para archivos
frame_archivos = tk.Frame(root, borderwidth=2, relief="groove",bg="white")
frame_archivos.grid(row=1, column=0, columnspan=4,padx=5, pady=(0,5))
# Paso 2
title_frame_archivos = tk.Label(root, text="Paso 2: Seleccione las tablas a crear", font=("Calibri", 12),bg="white", fg="#29669b")
title_frame_archivos.grid(row=2, column=0, columnspan=5,padx=5, sticky="w")
# Frame para contener las hojas
contenedor_hojas = tk.Frame(root, borderwidth=1, relief="groove")
contenedor_hojas.grid(row=3, column=0, columnspan=4,padx=5, pady=0,sticky="w")

# Frame para hoja 1
f_hoja1 = tk.Frame(contenedor_hojas, borderwidth=0, relief="groove")
f_hoja1.grid(row=0, column=0, padx=(0,5), pady=5, sticky="n")
title_frame_hoja1 = tk.Label(f_hoja1, text="1. Distritales", font=("Calibri", 11),fg="#000000")
title_frame_hoja1.grid(row=0, column=0, pady=5)

# Frame para hoja 2
f_hoja2 = tk.Frame(contenedor_hojas, borderwidth=0, relief="groove")
f_hoja2.grid(row=0, column=1, padx=(0,5), pady=5, sticky="n")
title_frame_hoja2 = tk.Label(f_hoja2, text="2. Análisis EAPB", font=("Calibri", 11))
title_frame_hoja2.grid(row=0, column=0, pady=5)

# Frame para hoja 3
f_hoja3 = tk.Frame(contenedor_hojas, borderwidth=0, relief="groove")
f_hoja3.grid(row=0, column=2, padx=(0,5), pady=5, sticky="n")
title_frame_hoja3 = tk.Label(f_hoja3, text="3. Suficiencia EAPB", font=("Calibri", 11))
title_frame_hoja3.grid(row=0, column=0, pady=5)

# Frame para hoja 4
f_hoja4 = tk.Frame(contenedor_hojas, borderwidth=0, relief="groove")
f_hoja4.grid(row=0, column=3, padx=(0,5), pady=5, sticky="n")
title_frame_hoja4 = tk.Label(f_hoja4, text="4. Suficiencia Distrital", font=("Calibri", 11))
title_frame_hoja4.grid(row=0, column=0, pady=5)

# frame para aviso varios
f_hint = tk.Frame(root, borderwidth=1, width=524,relief="groove", bg="white")
#f_hint.grid(row=4, column=0, columnspan=3, padx=(5,0), pady=5, sticky="nw")
f_hint.place(x=5, y=480, width=524, height=25)

# Selección de carpeta de origen de RIPSS
# Campo de captura de texto
label1 = tk.Label(frame_archivos, text="Archivo RIPSS depurados ",font=("Calibri", 10),bg="white")
label1.grid(row=1, column=0, pady=5, padx=(5,0), sticky="w")
entry = tk.Entry(frame_archivos, width=43)
entry.grid(row=1, column=1, columnspan=2, pady=5,sticky="w")
archivo_destino = os.path.join(os.environ['USERPROFILE'], 'Documents')
archivo_destino = recupera_parametro("archivo_destino")
entry.insert(0, archivo_destino)
boton_sel_carpeta = tk.Button(frame_archivos, image=icon_image_folder, command=lambda: selecciona_carpeta(1))
boton_sel_carpeta.grid(row=1, column=5, padx=(8,7), sticky="w")

label2 = tk.Label(frame_archivos, text="Archivo Prestadores ",font=("Calibri", 10),bg="white")
label2.grid(row=2, column=0, pady=5, padx=(5,0), sticky="w")
entry2 = tk.Entry(frame_archivos, width=43)
entry2.grid(row=2, column=1, columnspan=2, pady=5,sticky="w")
archivo_origen = os.path.join(os.environ['USERPROFILE'], 'Documents')
archivo_origen = recupera_parametro("archivo_origen")
entry2.insert(0, archivo_origen)
boton_sel_carpeta2 = tk.Button(frame_archivos, image=icon_image_folder, command=lambda: selecciona_carpeta(2))
boton_sel_carpeta2.grid(row=2, column=5, padx=(8,7), sticky="w")

label3 = tk.Label(frame_archivos, text="Archivo Códigos capacidad Instalada",font=("Calibri", 10),bg="white")
label3.grid(row=3, column=0, pady=5, padx=(5,0), sticky="w")
entry3 = tk.Entry(frame_archivos, width=43)
entry3.grid(row=3, column=1, columnspan=2, pady=5,sticky="w")
archivo_codigos_capacidad = os.path.join(os.environ['USERPROFILE'], 'Documents')
archivo_codigos_capacidad = recupera_parametro("archivo_codigos_capacidad")
entry3.insert(0, archivo_codigos_capacidad)
boton_sel_carpeta3 = tk.Button(frame_archivos, image=icon_image_folder, command=lambda: selecciona_carpeta(3))
boton_sel_carpeta3.grid(row=3, column=5, padx=(8,7), sticky="w")

checkbox_tabla1 = tk.IntVar()
checkbox_tabla2 = tk.IntVar()
checkbox_tabla3 = tk.IntVar()
checkbox_tabla4 = tk.IntVar()
checkbox_tabla5 = tk.IntVar()
checkbox_tabla6 = tk.IntVar()
checkbox_tabla7 = tk.IntVar()
checkbox_tabla8 = tk.IntVar()
checkbox_tabla9 = tk.IntVar()
checkbox_tabla10 = tk.IntVar()
checkbox_tabla11 = tk.IntVar()
checkbox_tabla12 = tk.IntVar()
checkbox_tabla13 = tk.IntVar()
checkbox_tabla14 = tk.IntVar()
checkbox_tabla15 = tk.IntVar()
checkbox_tabla16 = tk.IntVar()
checkbox_tabla17 = tk.IntVar()
checkbox_tabla18 = tk.IntVar()
checkbox_tabla19 = tk.IntVar()
checkbox_tabla20 = tk.IntVar()
checkbox_tabla21 = tk.IntVar()
checkbox_tabla22 = tk.IntVar()
checkbox_tabla23 = tk.IntVar()

tabla1 = tk.Checkbutton(f_hoja1, text="Construir tabla 1",font=("Calibri", 8), variable=checkbox_tabla1)
tabla1.grid(row=5, column=0, padx=5,pady=2, sticky="w")
tabla1.select()
tabla2 = tk.Checkbutton(f_hoja1, text="Construir tabla 2",font=("Calibri", 8), variable=checkbox_tabla2)
tabla2.grid(row=6, column=0, padx=5,pady=2, sticky="w")
tabla2.select()
tabla3 = tk.Checkbutton(f_hoja1, text="Construir tabla 3",font=("Calibri", 8), variable=checkbox_tabla3)
tabla3.grid(row=7, column=0, padx=5,pady=2, sticky="w")
tabla3.select()
tabla4 = tk.Checkbutton(f_hoja1, text="Construir tabla 4",font=("Calibri", 8), variable=checkbox_tabla4)
tabla4.grid(row=8, column=0, padx=5,pady=2, sticky="w")
tabla4.select()
tabla5 = tk.Checkbutton(f_hoja1, text="Construir tabla 5",font=("Calibri", 8), variable=checkbox_tabla5)
tabla5.grid(row=9, column=0, padx=5,pady=2, sticky="w")
tabla5.select()
tabla6 = tk.Checkbutton(f_hoja1, text="Construir tabla 6",font=("Calibri", 8), variable=checkbox_tabla6)
tabla6.grid(row=10, column=0, padx=5,pady=2, sticky="w")
tabla6.select()
tabla7 = tk.Checkbutton(f_hoja1, text="Construir tabla 7",font=("Calibri", 8), variable=checkbox_tabla7)
tabla7.grid(row=11, column=0, padx=5,pady=2, sticky="w")
tabla7.select()
tabla8 = tk.Checkbutton(f_hoja1, text="Construir tabla 8",font=("Calibri", 8), variable=checkbox_tabla8)
tabla8.grid(row=12, column=0, padx=5,pady=2, sticky="w")
tabla8.select()
tabla9 = tk.Checkbutton(f_hoja1, text="Construir tabla 9",font=("Calibri", 8), variable=checkbox_tabla9)
tabla9.grid(row=13, column=0, padx=5,pady=2, sticky="w")
tabla9.select()


tabla10 = tk.Checkbutton(f_hoja2, text="Construir tabla 10",font=("Calibri", 8), variable=checkbox_tabla10)
tabla10.grid(row=2, column=0, padx=5,pady=5)
tabla10.select()
tabla11 = tk.Checkbutton(f_hoja2, text="Construir tabla 11",font=("Calibri", 8), variable=checkbox_tabla11)
tabla11.grid(row=3, column=0, padx=5,pady=5)
tabla11.select()
tabla12 = tk.Checkbutton(f_hoja2, text="Construir tabla 12",font=("Calibri", 8), variable=checkbox_tabla12)
tabla12.grid(row=4, column=0, padx=5,pady=5)
tabla12.select()
tabla13 = tk.Checkbutton(f_hoja2, text="Construir tabla 13",font=("Calibri", 8), variable=checkbox_tabla13)
tabla13.grid(row=5, column=0, padx=5,pady=5)
tabla13.select()
tabla14 = tk.Checkbutton(f_hoja2, text="Construir tabla 14",font=("Calibri", 8), variable=checkbox_tabla14)
tabla14.grid(row=6, column=0, padx=5,pady=5)
tabla14.select()
tabla15 = tk.Checkbutton(f_hoja2, text="Construir tabla 15",font=("Calibri", 8), variable=checkbox_tabla15)
tabla15.grid(row=7, column=0, padx=5,pady=5)
tabla15.select()
tabla16 = tk.Checkbutton(f_hoja2, text="Construir tabla 16",font=("Calibri", 8), variable=checkbox_tabla16)
tabla16.grid(row=8, column=0, padx=5,pady=5)
tabla16.select()

tabla17 = tk.Checkbutton(f_hoja3, text="Construir tabla 17",font=("Calibri", 8),state="disabled", variable=checkbox_tabla17)
tabla17.grid(row=1, column=0, padx=5,pady=5)
tabla17.deselect()
tabla18 = tk.Checkbutton(f_hoja3, text="Construir tabla 18",font=("Calibri", 8),state="disabled", variable=checkbox_tabla18)
tabla18.grid(row=2, column=0, padx=5,pady=5)
tabla18.deselect()
tabla19 = tk.Checkbutton(f_hoja3, text="Construir tabla 19",font=("Calibri", 8),state="disabled", variable=checkbox_tabla19)
tabla19.grid(row=3, column=0, padx=5,pady=5)
tabla19.deselect()
tabla20 = tk.Checkbutton(f_hoja3, text="Construir tabla 20",font=("Calibri", 8),state="disabled", variable=checkbox_tabla20)
tabla20.grid(row=4, column=0, padx=5,pady=5)
tabla20.deselect()
tabla21 = tk.Checkbutton(f_hoja3, text="Construir tabla 21",font=("Calibri", 8),state="disabled", variable=checkbox_tabla21)
tabla21.grid(row=5, column=0, padx=5,pady=5)
tabla21.deselect()

tabla22 = tk.Checkbutton(f_hoja4, text="Construir tabla 22",font=("Calibri", 8),state="disabled", variable=checkbox_tabla22)
tabla22.grid(row=1, column=0, padx=5,pady=5)
tabla22.deselect()

# Selección de EPS
eps = listar_eps()
combobox_eps = ttk.Combobox(f_hoja2,width=15, font=("Calibri", 8),values=eps)
combobox_eps.grid(row=1, column=0, pady=(5,2))
combobox_eps.set(eps[0])

style = ttk.Style()
style.configure("Custom1.TButton", font=("Calibri", 10), background="white", foreground="#29669b")
style.map("Custom1.TButton", background=[("active", "#8795de"),("!pressed", "#FFFFFF"),("!active", "#FFFFFF")], foreground=[("active", "#29669b")])

boton_Hoja1 = ttk.Button(contenedor_hojas, style="Custom1.TButton",text=" Crear libro ", command=crea_libro)
boton_Hoja1.grid(row=1, column=0,columnspan=4,padx=5,pady=(0,5))

hint = tk.Label(f_hint, text="Seleccione los archivos para trabajar",font=("Calibri", 9),fg="#8795de",bg="white")
hint.grid(row=0, column=0, columnspan=2, padx=(5,0))
entry.bind("<FocusOut>", lambda event: actualizar_combobox())
# Actualiza programa
update_program()
# Mostrar la ventana
root.mainloop()