import time
import pandas as pd
import openpyxl
import os
import sys
import requests
import shutil
import re
import lxml
import subprocess
import io
import tkinter as tk
import socket
import threading
import platform
import psutil

from tqdm import tqdm
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.firefox.options import Options
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from tkinter import PhotoImage
from tkinter import ttk
from tkinter import filedialog
from urllib.parse import urljoin

# === Parametros generales
version = "1.0"
download_url = "https://misejecutables.000webhostapp.com" # Ubicación de nuevas versiones
# === Funciones para auto-actualizar el programa
def copia_nueva_version():
    # Si el archivo que se ejecuta se llama nueva_version, copia este archivo en baja_archivos
    if os.path.basename(sys.argv[0]) == "nueva_version.exe":
        try:
            os.replace("nueva_version.exe", "baja_archivos.exe")
        except Exception as e:
            print("Error al reemplazar el archivo:", e)
            sys.exit(1)
        print("¡Actualización completada! Reiniciando el programa...")
        # Reiniciar el programa con el nombre original
        os.system("start baja_archivos.exe")
        sys.exit()


def download_latest_version(url, filename):
    url = urljoin(url, filename)
    try:
        response = requests.get(url)
        print(response.text)
        return response.text
        with open(filename, 'wb') as f:
            f.write(response.content)
    except Exception as e:
        return 0

def restart_program():
    python = sys.executable
    os.execl(python, python, *sys.argv)

def extrae_ultima_version(url, filename):
    new_filename = "nueva_version.exe"
    with open(new_filename, 'wb') as f:
        url = urljoin(url, filename)
        print(url)
        response = requests.get(url, stream=True)
        total_length = response.headers.get('content-length')
        print(total_length)
        if total_length is None:
            f.write(response.content)
        else:
            dl = 0
            total_length = int(total_length)
            print("Directorio de trabajo actual:", os.getcwd())
            progress_bar = tqdm(total=total_length, unit='B', unit_scale=True, ncols=80, miniters=1)
            for data in response.iter_content(chunk_size=4096):
                dl += len(data)
                f.write(data)
                progress_bar.update(len(data))
            progress_bar.close()
def boton_actualizar():
    boton_actualiza = tk.Button(root, text="Actualizar versiónn", command=actualiza)
    boton_actualiza.grid(row=4, column=1,columnspan=1, pady=1)
def update_program():
    global version, download_url
    current_version = version  # Obtener la versión actual de tu programa
    last_version = download_latest_version(download_url,"version.txt")
    print("Version actual: "+ current_version + " Ultima versión: " + last_version)
    if current_version < last_version:
        boton_actualizar()
        return
    else:
        print("No hay actualizaciones disponibles.")
def actualiza():
        # Descargar el nuevo ejecutable
        global download_url
        new_executable = "baja_archivos.exe"
        extrae_ultima_version(download_url, new_executable)

        # Cerrar el archivo en ejecución ("baja_archivos.exe")
        # Reiniciar el programa con el nombre original
        print("¡Archivo descargao! Reiniciando el programa...")
        os.system("start nueva_version.exe")
        sys.exit()

# === Fin funciones de auto-actualización
def listar_archivos_R(carpeta):
    archivos_r = [archivo for archivo in os.listdir(carpeta) if archivo.endswith('.R')]
    return archivos_r
def copia_archivo(origen, destino):
    # Toma el archivo de download mas nuevo y lo copia destino quitandole el (##) del final
    # Obtener una lista de todos los archivos en la carpeta "origen"
    archivos_origen = os.listdir(origen)

    # Filtrar la lista para obtener solo archivos (excluir directorios)
    archivos_origen = [archivo for archivo in archivos_origen if os.path.isfile(os.path.join(origen, archivo))]

    # Ordenar la lista por fecha de modificación (más reciente primero)
    archivos_origen.sort(key=lambda x: os.path.getmtime(os.path.join(origen, x)), reverse=True)

    # Tomar el último archivo de la lista (el más reciente)
    if archivos_origen:
        ultimo_archivo = archivos_origen[0]

        # Ruta completa del último archivo en "origen"
        ruta_ultimo_archivo = os.path.join(origen, ultimo_archivo)
        # Extraer el nombre del archivo y la extensión
        nombre_archivo, extension = os.path.splitext(ultimo_archivo)

        # Eliminar los números al final del nombre original del archivo (si los hay)
        nuevo_nombre_archivo = re.sub(r'\(\d+\)$', '', nombre_archivo)
        print("Nuevo Nombre: "+nuevo_nombre_archivo)
        # Concatenar el nuevo nombre con la extensión
        nuevo_nombre_completo = nuevo_nombre_archivo + extension

        # Ruta completa del último archivo en "origen"
        ruta_ultimo_archivo = os.path.join(origen, ultimo_archivo)

        # Ruta completa del nuevo archivo en "destino"
        ruta_nuevo_archivo = os.path.join(destino, nuevo_nombre_completo)
        # Copiar el archivo a la carpeta "destino"
        print("Lin 132 "+ruta_nuevo_archivo)
        
        shutil.copy(ruta_ultimo_archivo, ruta_nuevo_archivo)
        salida = os.path.join(destino, nuevo_nombre_archivo + ".xlsx")
        convertir_html_a_excel(ruta_nuevo_archivo, salida)
        print(f"El último archivo '{ultimo_archivo}' ha sido copiado a la carpeta '{destino}'.")
    else:
        print(f"La carpeta '{origen}' está vacía. No se encontraron archivos para copiar.")

def convertir_html_a_excel(ruta_archivo_entrada, ruta_archivo_salida):
    try:
        # Leer el archivo HTML y obtener la lista de DataFrames
        dataframes = pd.read_html(ruta_archivo_entrada,header=0)
        
        if not dataframes:
            print(f"No se encontraron tablas en el archivo HTML '{ruta_archivo_entrada}'.")
            return

        # Obtener el primer DataFrame de la lista
        df = dataframes[0]
        # Obtiene el nombre del archivo
        ruta_archivo = Path(ruta_archivo_salida)
        nombre_archivo_completo = ruta_archivo.name
        nombre_archivo, extension = os.path.splitext(nombre_archivo_completo)
        print("Nombre archivo: "+nombre_archivo)
        # Guardar el DataFrame en un archivo Excel
        df.to_excel(ruta_archivo_salida,  sheet_name=nombre_archivo, index=False, startrow=3)

        print(f"Archivo Excel '{ruta_archivo_salida}' creado exitosamente.")
    except Exception as e:
        print(f"Se produjo un error al convertir el archivo HTML a Excel: {e}")
#***
def convertir_html_a_excel_dos(ruta_archivo_entrada, ruta_archivo_salida):
    try:
        # Leer el archivo HTML y obtener la lista de DataFrames
        dataframes = pd.read_html(ruta_archivo_entrada, header=0)
        
        if not dataframes:
            print(f"No se encontraron tablas en el archivo HTML '{ruta_archivo_entrada}'.")
            return

        # Obtener el primer DataFrame de la lista
        df = dataframes[0]

        # Crear un nuevo archivo Excel
        wb = Workbook()
        ws = wb.active

        # Parsear el HTML para obtener el encabezado con el formato original
        with open(ruta_archivo_entrada, "r", encoding="ISO-8859-1") as archivo_html:
            contenido_html = archivo_html.read()
            soup = BeautifulSoup(contenido_html, "html.parser")
            tabla_html = soup.find("table")
            if tabla_html:
                filas_html = tabla_html.find_all("tr")
                filas_html_br = tabla_html.find_all("br")
                if filas_html:
                    num_filas_html = len(filas_html) + len(filas_html_br)
                    num_filas_excel = ws.max_row
                    fila_inicio_tabla = num_filas_excel + 2 if num_filas_excel < num_filas_html else 1

                    # Buscar el texto que está antes de la tabla en el HTML
                    texto_antes_de_tabla = contenido_html.split(str(tabla_html))[0].strip()
                    # Eliminar caracteres especiales y espacios en blanco adicionales
                    texto_antes_de_tabla = re.sub(r'\s+', ' ', texto_antes_de_tabla)
                    soup = BeautifulSoup(texto_antes_de_tabla, "html.parser")
                    texto_antes_de_tabla = soup.get_text()
                    # Escribir el texto antes de la tabla en una celda de Excel
                    ws.cell(row=fila_inicio_tabla - 1, column=1, value=texto_antes_de_tabla)

                    for i, fila_html in enumerate(filas_html):
                        celdas_html = fila_html.find_all(["th", "td"])
                        for j, celda_html in enumerate(celdas_html):
                            valor_celda = celda_html.get_text()
                            celda_excel = ws.cell(row=fila_inicio_tabla + i, column=j+1, value=valor_celda)

                            # Copiar el formato de celda del HTML al Excel
                            if celda_html.get("bgcolor"):
                                fill_color = celda_html["bgcolor"]
                                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                                celda_excel.fill = fill

        # Escribir la tabla en Excel usando openpyxl
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        # Obtener el nombre del archivo sin extensión
        ruta_archivo = Path(ruta_archivo_salida)
        nombre_archivo_completo = ruta_archivo.name
        nombre_archivo, extension = os.path.splitext(nombre_archivo_completo)

        # Cambiar el nombre de la hoja en el archivo Excel
        ws.title = nombre_archivo

        # Guardar el archivo Excel
        wb.save(ruta_archivo_salida)

        print(f"Archivo Excel '{ruta_archivo_salida}' creado exitosamente.")
    except Exception as e:
        print(f"Se produjo un error al convertir el archivo HTML a Excel: {e}")

#***

def baja_capacidadInstalada():
    global ruta_downloads, headless
    global ruta_archivos_recibidos
    # Crea nuevo driver
    #driver = webdriver.Edge()
    headless = False
    headless = recupera_parametro("headless")
    if headless:
        firefox_options = Options()
        firefox_options.add_argument("--headless")
    else:
        firefox_options = ""
    driver = webdriver.Firefox(options=firefox_options)

    # Navigate al REPS

    driver.get("https://prestadores.minsalud.gov.co/habilitacion/")

    driver.switch_to.frame("areawork")
    time.sleep(2)
    button_modal = driver.find_element(by="xpath", value="//button[@class='btn btn-secondary']")
    button_modal.click()
    time.sleep(2)
    button_login = driver.find_element(by="xpath", value="//*[@id='Button1']")
    button_login.click()
    time.sleep(2)
    link_registro_actual = driver.find_element(by="xpath", value='/html/body/form/table/tbody/tr[3]/td/table/tbody/tr[4]/td/span/table/tbody/tr/td[1]/table[2]/tbody/tr[2]/td/div/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/a')
    link_registro_actual.click()
    time.sleep(2)
    handles = driver.window_handles

    # Switch to the second tab
    driver.switch_to.window(handles[1])
    boton_capacidad = driver.find_element(by="xpath", value="//*[@id='_ctl0_ContentPlaceHolder1_btn_capacidad_reps']")
    boton_capacidad.click()
    time.sleep(2)

    select = driver.find_element(by="xpath", value='//*[@id="_ctl0_ContentPlaceHolder1_ddsede_departamento"]')

    select_departamento = Select(select)
    select_departamento.select_by_value('11')
    select = driver.find_element(by="xpath", value='//*[@id="_ctl0_ContentPlaceHolder1_ddsede_municipio"]')
    select_municipio = Select(select)
    select_municipio.select_by_value('11001')
    select = driver.find_element(by="xpath", value='//*[@id="_ctl0_ContentPlaceHolder1_ddgrupo_capacidad"]')
    select_grupo = Select(select)
    select_grupo.select_by_value('CAMAS')
    boton_buscar = driver.find_element(by="xpath", value='//*[@id="_ctl0_ibBuscarFtr"]')
    boton_buscar.click()
    time.sleep(2)
    boton_excel = driver.find_element(by="xpath", value='//*[@id="_ctl0_ContentPlaceHolder1_ibExcel"]')
    boton_excel.click()
    time.sleep(10)

    # Llamar a la función copia_archivo con las rutas especificadas
    copia_archivo(ruta_downloads, ruta_archivos_recibidos)

    # Close the driver
    driver.quit()

def baja_sirc():
    global ruta_downloads, headless
    global ruta_archivos_recibidos
    headless = False
    headless = recupera_parametro("headless")
    if headless:
        firefox_options = Options()
        firefox_options.add_argument("--headless")
    else:
        firefox_options = ""
    driver = webdriver.Firefox(options=firefox_options)

    # Navega al SIRC
    driver.get("http://app.saludcapital.gov.co/sirc2/")
    link_ingresar = driver.find_element(by="xpath", value='//*[@id="dnn_dnnLOGIN_cmdLogin"]')
    link_ingresar.click()
    time.sleep(2)
    usuario =  driver.find_element(by="xpath", value='//*[@id="dnn_ctr_Login_Login_DNN_txtUsername"]')
    password = driver.find_element(by="xpath", value= '//*[@id="dnn_ctr_Login_Login_DNN_txtPassword"]')
    usuario.send_keys("JePTriana")
    password.send_keys("JePTriana2021+")
    btn_login = driver.find_element(by="xpath", value='//*[@id="dnn_ctr_Login_Login_DNN_cmdLogin"]')
    btn_login.click()
    time.sleep(2)
    menu_modulos = driver.find_element(by="xpath", value='//*[@id="dnn_dnnNAV_ctldnnNAVt57"]')
    menu_modulos.click()
    time.sleep(1)
    item_hospitalizacion = driver.find_element(by="xpath", value='//*[@id="dnn_dnnNAV_ctldnnNAVt316"]')
    item_hospitalizacion.click()
    time.sleep(1)
    item_capacidad = driver.find_element(by="xpath", value='//*[@id="dnn_dnnNAV_ctldnnNAVt253"]')
    item_capacidad.click()
    time.sleep(1)
    item_ocupacion = driver.find_element(by="xpath", value='//*[@id="dnn_dnnNAV_ctldnnNAVt303"]')
    item_ocupacion.click()
    time.sleep(1)
    item_salida = driver.find_element(by="xpath", value='//*[@id="dnn_ctr723_dnnACTIONBUTTON1_lnk2"]')
    item_salida.click()
    time.sleep(7)
    select = driver.find_element(by="xpath", value='//*[@id="dnn_ctr723_ReporteCensoCamas_ddlEntidad"]')
    select_entidad = Select(select)
    select_entidad.select_by_value('Todas')
    time.sleep(2)
    select = driver.find_element(by="xpath", value='//*[@id="dnn_ctr723_ReporteCensoCamas_ddlTipoEntidad"]')
    select_tipo = Select(select)
    select_tipo.select_by_value('Todos')
    select = driver.find_element(by="xpath", value='//*[@id="dnn_ctr723_ReporteCensoCamas_DDList_Grupos"]')
    select_grupos = Select(select)
    select_grupos.select_by_value('Todos')
    fecha_inicio = driver.find_element(by="xpath", value='//*[@id="dnn_ctr723_ReporteCensoCamas_txtFechaInicial"]')
    fecha_fin = driver.find_element(by="xpath", value='//*[@id="dnn_ctr723_ReporteCensoCamas_txtFechaFinal"]')
    fecha_actual = datetime.now()
    fecha_formateada = fecha_actual.strftime("%d/%m/%Y")
    fecha_inicio.send_keys(fecha_formateada)
    fecha_fin.send_keys(fecha_formateada)
    time.sleep(5)
    btn_consultar = driver.find_element(by="xpath", value='//*[@id="dnn_ctr723_ReporteCensoCamas_btnConsultar"]')
    btn_consultar.click()
    time.sleep(15)
    lnk_exportar = driver.find_element(by="xpath", value='//*[@id="dnn_ctr723_ReporteCensoCamas_LinkButton4"]')
    lnk_exportar.click()
    copia_archivo(ruta_downloads, ruta_archivos_recibidos)
    # Close the driver
    driver.quit()

def ejecuta_programa_R(ruta_programa_r):
    # Verifica archivo abierto
    global ruta_archivos_recibidos
    abierto = False
    abierto = archivo_esta_abierto(ruta_archivos_recibidos + "/Faltan.xlsx")
    print(ruta_archivos_recibidos + "/Faltan.xlsx")
    cerrar_libro_excel(os.path.join(ruta_archivos_recibidos, "Faltan.xlsx"))
    if abierto:
        hint.config(text="Cierre archivo de Excel y vuelva a ejecutar")
        hint.update()
    else:
        # Ejecutar el programa de R
        try:
            comando = f"Rscript \"{ruta_programa_r}\""
            resultado = subprocess.run(comando, shell=True, check=True)
            if resultado.returncode == 0:
                print("Programa "+ruta_programa_r+" ha corrido exitosamente")
            else:
                print("Error al ejecutar el programa de R:")
                print("Salida estándar:")
                print(resultado.stdout)
                print("Error estándar:")
                print(resultado.stderr)
        except FileNotFoundError:
            print(f"El archivo '{ruta_programa_r}' no fue encontrado.")
        except Exception as e:
            print(f"Se produjo un error al leer el archivo: {e}")
            raise e 
def archivo_esta_abierto(ruta_archivo):
    try:
        with open(ruta_archivo, "w"):
            return False  # El archivo no está abierto por otro proceso
    except IOError:
        return True  # El archivo está abierto por otro proceso

def hay_conexion_internet():
    try:
        # Intentar conectarse a un servidor de ejemplo (puedes cambiar el dominio y el puerto)
        socket.create_connection(("prestadores.minsalud.gov.co", 80))
        return True  # Hay conexión a Internet
    except OSError:
        return False  # No hay conexión a Internet

   
def abre_archivo_excel(ruta_archivo_excel):
    # cerrar_libro_excel(ruta_archivo_excel)
    # Comando para abrir el archivo con Excel en el sistema operativo
    comando = f"start excel \"{ruta_archivo_excel}\""

    # Ejecutar el comando para abrir el archivo con Excel
    try:
        subprocess.run(comando, shell=True, check=True)
        print(f"El archivo '{ruta_archivo_excel}' se abrió exitosamente en Excel.")
    except subprocess.CalledProcessError as e:
        print(f"Se produjo un error al intentar abrir el archivo en Excel: {e}")

def asigna_parametro(variable):
    # Nombre del archivo config.txt
    file_name = "config.txt"
    
    try:
        # Intentamos abrir el archivo en modo lectura
        with open(file_name, "r") as file:
            lines = file.readlines()
    except FileNotFoundError:
        # Si el archivo no existe, lo creamos y escribimos el parámetro con su valor
        with open(file_name, "w") as file:
            file.write(f"{variable}={globals()[variable]}\n")
    else:
        # Buscamos la variable en las líneas del archivo
        for i, line in enumerate(lines):
            key, value = line.strip().split("=")
            if key == variable:
                # Cambiamos el valor de la variable en el archivo
                lines[i] = f"{variable}={globals()[variable]}\n"
                with open(file_name, "w") as file:
                    file.writelines(lines)
                break
        else:
            # Si la variable no está en el archivo, la agregamos al final
            with open(file_name, "a") as file:
                file.write(f"{variable}={globals()[variable]}\n")

def recupera_parametro(variable):
    # Nombre del archivo config.txt
    file_name = "config.txt"
    
    try:
        # Intentamos abrir el archivo en modo lectura
        with open(file_name, "r") as file:
            lines = file.readlines()
    except FileNotFoundError:
        # Si el archivo no existe, devolvemos el valor original de la variable
        return globals()[variable]
    else:
        # Buscamos la variable en las líneas del archivo
        for line in lines:
            key, value = line.strip().split("=")
            if key == variable:
                return value
        # Si la variable no está en el archivo, devolvemos el valor original de la variable
        return globals()[variable]


#-*-*-*-*
# Función para ejecutar los pasos seleccionados por el usuario
def pasos_ejecutar():
    global ruta_archivos_recibidos, nombre_programa
    ruta_archivos_recibidos = entry.get()
    asigna_parametro("ruta_archivos_recibidos")
    nombre_programa = combobox.get()
    asigna_parametro("nombre_programa")
    
    if checkbox_reps.get():
        hint.config(text="Programa en ejecución")
        hint.update()
        baja_capacidadInstalada()
        hint.config(text="")
        hint.update()
    if checkbox_sirc.get():
        hint.config(text="Programa en ejecución")
        hint.update()
        baja_sirc()
        time.sleep(5)
        hint.config(text="")
        hint.update()
    if checkbox_progr.get():
        hint.config(text="Programa en ejecución")
        hint.update()
        ejecuta_programa_R(ruta_archivos_recibidos + '/' + nombre_programa)
        time.sleep(8)
        abre_archivo_excel(os.path.join(ruta_archivos_recibidos,"Faltan.xlsx"))
        hint.config(text="")
        hint.update()
def keep_on_top():
    root.lift()
    root.update()
    root.after(1000, keep_on_top)  # Ejecutar la función cada 1000 milisegundos
def informa_conexion_internet():
    conexion = hay_conexion_internet()
    if not conexion:
        hint2.config(text="No hay conexión a Internet")
        icon_label.config(image=icon_image_disconnect)
    else:
        hint2.config(text="Conectado")
        icon_label.config(image=icon_image)
    icon_label.grid(row=8, column=3)
    root.after(1000, informa_conexion_internet)  # Llamar a la función nuevamente después de 1000 milisegundos (1 segundo)
def selecciona_carpeta():
    global ruta_archivos_recibidos
    ruta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta")
    if ruta_seleccionada:
        entry.delete(0, tk.END)  # Limpiar el contenido actual del Entry
        entry.insert(0, ruta_seleccionada)
        ruta_archivos_recibidos = ruta_seleccionada
        actualizar_combobox()
def actualizar_combobox():
    archivos_r = listar_archivos_R(entry.get())
    combobox['values'] = archivos_r
    if len(archivos_r) > 0:
        combobox.set(archivos_r[0])
    else:
        combobox.set("")
def cerrar_libro_excel(ruta_archivo):
    nombre = "Faltan.xlsx"
    for proceso in psutil.process_iter(attrs=['pid', 'name']):
        if proceso.info['name'] == nombre:
            pid = proceso.info['pid']
            try:
                p = psutil.Process(pid)
                archivos_abiertos = p.open_files()
            except Exception as e:
                print(f"Error al procesar {nombre}: {e}")
            for archivo in archivos_abiertos:
                if archivo.path == os.path.normpath(ruta_archivo):
                    print(f"Archivo {ruta_archivo} abierto por el proceso {pid}")
                    try:
                        p.terminate()  # Termina el proceso que tiene el archivo abierto
                        print(f"Proceso {nombre} (PID: {pid}) cerrado.")
                    except Exception as e:
                        print(f"Error al cerrar el proceso {nombre}: {e}")


#PROGRAMA PRINCIPAL

print("Nombre del ejecutable: "+os.path.basename(sys.argv[0]))
copia_nueva_version() # En caso de actualización
script_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
images_dir = os.path.join(script_dir, 'images')
# Ruta de la carpeta "C:/Archivos recibidos"
ruta_archivos_recibidos = "C:/Archivos recibidos"
ruta_archivos_recibidos = recupera_parametro("ruta_archivos_recibidos")
#Programa a correr en R pro omisión
nombre_programa = "Reportes_fal_dif.R"
nombre_programa = recupera_parametro("nombre_programa")
# Ruta de la carpeta "Downloads"
user_home = os.environ["USERPROFILE"]
download_dir = os.path.join(user_home, "Downloads")
print("Ruta download: "+download_dir)
ruta_downloads = download_dir

# Ruta de la carpeta "Downloads"
user_home = os.environ["USERPROFILE"]
download_dir = os.path.join(user_home, "Downloads")
print("Ruta download: "+download_dir)
ruta_downloads = download_dir

# Configuración de la ventana
root = tk.Tk()
root.title("Reportes diarios de capacidad (camas)  Ver.("+version+")")
root.geometry("480x220")
logo_path = os.path.join(images_dir, 'Logo.ico')
root.iconbitmap(logo_path) 
root.lift()
keep_on_top()
# Simbolos de conexión y carpeta

icon_path = os.path.join(images_dir, 'wifi.png')
icon_image = PhotoImage(file=icon_path)
icon_path1 = os.path.join(images_dir, 'nowifi.png')
icon_image_disconnect = PhotoImage(file=icon_path1)
icon_label = tk.Label(root, image=icon_image)
icon_path = os.path.join(images_dir, 'folder.png')
icon_image_folder = PhotoImage(file=icon_path)
# Etiqueta para indicar al usuario que ingrese texto
etiqueta = tk.Label(root, text="    Ubicación de los programas en R:")
etiqueta.grid(row=2, column=1, pady=10)

# Campo de captura de texto
entry = tk.Entry(root, width=30)
entry.grid(row=2, column=2, pady=10,sticky="e")
entry.insert(0, ruta_archivos_recibidos)
boton_sel_carpeta = tk.Button(root, image=icon_image_folder, command=selecciona_carpeta)
boton_sel_carpeta.grid(row=2, column=3,sticky="w")

# Variable para almacenar el estado del checkbox
checkbox_reps = tk.IntVar()
checkbox_sirc = tk.IntVar()
checkbox_progr = tk.IntVar()

# Crear los checkboxes en diferentes filas
checkbox1 = tk.Checkbutton(root, text="1. Descargar Capacidad Instalada", variable=checkbox_reps)
checkbox1.grid(row=3, column=2, padx=5, sticky="w")
checkbox1.select()

checkbox2 = tk.Checkbutton(root, text="2. Descargar Porcentaje Ocupación", variable=checkbox_sirc)
checkbox2.grid(row=4, column=2, padx=5, sticky="w")
checkbox2.select()

checkbox3 = tk.Checkbutton(root, text="3. Correr programa en R", variable=checkbox_progr)
checkbox3.grid(row=5, column=2, padx=5, sticky="w")
checkbox3.select()

# Selección del programa R
archivos_r = listar_archivos_R(entry.get())
combobox = ttk.Combobox(root, values=archivos_r)
combobox.grid(row=6, column=2, pady=5)
combobox.set(nombre_programa)

# Avisos
hint = tk.Label(root, text="",  fg="green")
hint.grid(row=3, column=1, pady=5)
hint1 = tk.Label(root, text=" Debe tener el navegador Firefox Instalado",  fg="green")
hint1.grid(row=8, column=0,columnspan=2, pady=10)
hint2 = tk.Label(root, text="",  fg="green")
hint2.grid(row=8, column=2, columnspan=1, pady=10, sticky="e")


# Botón para ejecutar los pasos seleccionados
boton_obtener = tk.Button(root, text="Ejecutar", command=pasos_ejecutar)
boton_obtener.grid(row=7, column=2,columnspan=1, pady=1)

#Verifica conexión a Internet
informa_conexion_internet()
update_program()

entry.bind("<FocusOut>", lambda event: actualizar_combobox())
# Mostrar la ventana
root.mainloop()
