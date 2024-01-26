# Aplicación Para generar datos para el Visor de RIAS
# Toma servicios de Reps y el archivo Ruta-Servicio y genera 
# la hoja de datos del Visor
import datetime
# import time
import panel as pn
import tkinter as tk
# import subprocess
from tkinter import PhotoImage
from tkinter import ttk
from tkinter import filedialog
import threading
import itertools

#from tqdm import tqdm

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
#from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
#from urllib.parse import urljoin
import os
import sys
import json 
import re

# variables globales
version = "1.0"
ubicacion_fuentes = "ubicacion_fuentes.json"
objetos_tkinter = {}
contenido = {} # Diccionario con el contenido de los libros cargados, el indice es la clave
# Carpetas
script_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
images_dir = os.path.join(script_dir, 'images')
# Definicion de imagenes
logo_path = os.path.join(images_dir, 'Logo.ico')
icon_folder_path = os.path.join(images_dir, 'folder.png')

def draw_main_windows(width, height, title="Genera Data para Visor. version: " + version):
    global root
    root = tk.Tk()
    root.title(title)
    root.geometry(str(width) + "x" + str(height))
    root.configure(bg="white")
    root.resizable(False, False)
    root.iconbitmap(logo_path)
    hint_frame(width, height)

def hint_frame(x,y):
    global hint
    f_hint = tk.Frame(root, borderwidth=1, width=524,relief="groove", bg="white")
    f_hint.place(x=5, y=y-30, width=x-11, height=25)
    hint = tk.Label(f_hint, text="Listo",font=("Calibri", 9),fg="#8795de",bg="white")
    hint.grid(row=0, column=0, columnspan=2, padx=(5,0))
    
def selecciona_carpeta(indice,entry):
    directorio_inicial = ""
    ruta_seleccionada = filedialog.askopenfilename(title="Seleccione la fuente de Datos", 
                                                   initialdir=directorio_inicial,
                                                   filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")) ) 
    entry.delete(0, tk.END)
    entry.insert(0, ruta_seleccionada)
    # Registra objeto tkinter
    objetos_tkinter[str(indice)] = entry
    # Actualiza archivo config con información de las selecciones de archivos
    update_ubicacion(str(indice), ruta_seleccionada)
def update_ubicacion(clave, valor):
    # Lee archivo
    if os.path.exists(ubicacion_fuentes):
        with open(ubicacion_fuentes, "r") as archivo:
            loaded_data = json.load(archivo)
    else:
        loaded_data = {}
    loaded_data[clave] = valor
    # Graba archivo actualizado
    with open(ubicacion_fuentes, "w") as archivo:
        json.dump(loaded_data, archivo)
def get_ubicación(indice):
    if os.path.exists(ubicacion_fuentes):
        with open(ubicacion_fuentes, "r") as archivo:
            loaded_data = json.load(archivo)
    else:
        loaded_data = {}
    return loaded_data.get(str(indice), False)

def draw_files_frame(row,col,label="Seleccione archivo: "):
    global root
    global icon_folder_path
    
    frame_archivos = tk.Frame(root, borderwidth=2, relief="groove",bg="white")
    frame_archivos.grid(row=row, column=col, columnspan=4,padx=5, pady=(0,5), sticky="w")
    label1 = tk.Label(frame_archivos, text=label,font=("Calibri", 10),bg="white", width=20, anchor="w")
    label1.grid(row=row, column=col, pady=5, padx=(5,0), sticky="w")
    texto = get_ubicación(row)
    entry = tk.Entry(frame_archivos, width=43)
    entry.grid(row=row, column=col+1, columnspan=2, pady=5,sticky="w")
    if texto:
        entry.delete(0, tk.END)
        entry.insert(0, texto)
        # Registra objeto tkinter
        objetos_tkinter[str(row)] = entry
    icon_folder = tk.PhotoImage(file=icon_folder_path)
    boton_sel_carpeta = ttk.Button(frame_archivos, text="Abrir",image=icon_folder, command=lambda: selecciona_carpeta(row,entry))
    boton_sel_carpeta.grid(row=row, column=col+4, padx=(8,7), sticky="w")

def menu(row, col):
    frame_menu = tk.Frame(root, borderwidth=2, relief="groove",bg="white")
    frame_menu.grid(row=row, column=col, columnspan=4,padx=5, pady=(0,5))
    boton_carga_libros = ttk.Button(frame_menu, text="Cargar Archivos", command=lambda: carga_libros())
    boton_carga_libros.grid(row=row, column= col, padx=1)
    boton_lista_hojas = ttk.Button(frame_menu, text="Lista Hojas", command=lambda: lista_hojas())
    boton_lista_hojas.grid(row=row, column= col + 1, padx=1)
    boton_genera_datos = ttk.Button(frame_menu, text="Genera Datos", command=lambda: genera_datos())
    boton_genera_datos.grid(row=row, column= col + 2, padx=1)

def carga_libros():
    hint.config(text="Cargando libros ...")
    root.update()
    print("Objetos: ", len(objetos_tkinter))
    if len(objetos_tkinter) > 0:
        for indice in objetos_tkinter.keys():
            archivo = objetos_tkinter[indice].get()
            contenido[indice] = load_workbook(filename=archivo, read_only=True)
                
    hint.config(text="Libros cargados")
    root.update()
def lista_hojas():
    print("Listado de Hojas")
    for libro in contenido:
        nombre_hojas = contenido[libro].sheetnames
        for title in nombre_hojas:
            print(libro, re.sub(r'\s+\(\d+\)\s*$', '', title))
def genera_datos():
    hint.config(text="En ejecución ...")
    hint.update()
    libro_servicios = contenido['1']
    libro_rutas = contenido['2']
    # libro_visor = contenido['3']
    hoja_servicios = libro_servicios['Servicios (9)']
    hoja_rutas = libro_rutas['Hoja1']
    # hoja_datos_visor = libro_visor['DATOS']
    # Diccionario servicios en ruta
    # Recibe un servicio y devuelve una lista con las rutas que lo contienen
    diccionario_rutas = {}
    for row in hoja_rutas.iter_rows(min_row=2, values_only=True):
            servicio_ruta = str(row[0]).strip()
            ruta = str(row[1]).strip()
            ruta_nombre = str.upper(row[2]).strip()
            codigo = servicio_ruta.split("-")
            rutas = diccionario_rutas.get(codigo[0], [])
            rutas.append((ruta, ruta_nombre))
            diccionario_rutas[codigo[0]] = rutas
    
    j = 0
    tabla = [['naju_nombre', 'complejidades', 'zona', 'SERVICIO', 'RUTA', 'RUT_NOM', 'TIPO_DE_NIVEL']]
    for row in hoja_servicios.iter_rows(min_row=2, values_only=True):
        naju_nombre = str(row[14]).strip()
        complejidades = str(row[63]).strip()
        zona = str(row[89]).strip()
        cod_servicio = str(row[23]).strip()
        servicio = cod_servicio + '-' + str.upper(row[24]).strip()
        tipo_nivel = str(row[101]).strip()
        if tipo_nivel == 'Complementario':
            tipo_nivel = 'COMPLEMENTARIA'
        if tipo_nivel == 'Primario':
            tipo_nivel = 'PRIMARIO'
        if tipo_nivel == 'Sin Complejidad':
            tipo_nivel = 'SC'
        rutas = diccionario_rutas.get(cod_servicio, [])
        for ruta in rutas:
            rias = ruta[0]
            ruta_nombre = ruta[1]
            registro = [naju_nombre, complejidades, zona, servicio, rias, ruta_nombre, tipo_nivel]
            tabla.append(registro)
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'DATOS'
    
    # Ordena por servicio
    tabla[1:] = sorted(tabla[1:], key=lambda x: x[3])
    # Llena con nueva información
    for row in tabla:
        sheet.append(row)
    i = 0
    nombre_guardar = "Datos para visor.xlsx"
    while os.path.exists(nombre_guardar):
        i += 1
        nombre_guardar = 'Datos_para_visor' +" ("+str(i)+")"+".xlsx"
    wb.save(nombre_guardar)
    hint.config(text="Se ha generado el archivo: " + os.path.abspath(nombre_guardar))
     
# Programa principal
draw_main_windows(535,540)
draw_files_frame(1,0,"Servicios del REPS")
draw_files_frame(2,0,"Ruta-Servicio")
# draw_files_frame(3,0,"Visor")
menu(4, 0)

root.mainloop()